import os, time
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# === CONFIGURACIÓN DE RUTAS ===
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)  # crea la carpeta si no existe

# Una marca de tiempo única por ejecución (para Excel)
HORA_EXCEL = datetime.now().strftime("%H%M%S")


def log(mensaje, archivo_log):
    """Escribe un mensaje en el log TXT abierto."""
    hora = time.strftime("%Y-%m-%d %H:%M:%S")
    archivo_log.write(f"[{hora}] {mensaje}\n")
    archivo_log.flush()


def log_excel(nombre, mensaje="", negrita=False, color_mensaje="black"):
    """Escribe un registro en Excel dentro de /reports."""
    ahora = datetime.now()
    fecha = ahora.strftime("%Y-%m-%d")
    hora = ahora.strftime("%H:%M:%S")

    archivo_excel = os.path.join(REPORTS_DIR, f"registro_pruebas_{fecha}_{HORA_EXCEL}.xlsx")

    # Crear archivo nuevo si no existe
    if not os.path.exists(archivo_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro"
        ws.append(["Fecha", "Hora", "Nombre", "Mensaje"])
        wb.save(archivo_excel)

    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    ws.append([fecha, hora, nombre, mensaje])

    fila_actual = ws.max_row

    # === Estilos ===
    if negrita:
        font_bold = Font(bold=True)
        fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws[f"C{fila_actual}"].font = font_bold
        ws[f"C{fila_actual}"].fill = fill_gray

    center_alignment = Alignment(horizontal="center", vertical="center")

    if color_mensaje.lower() == "verde":
        ws[f"D{fila_actual}"].fill = PatternFill(start_color="B2E7D7", end_color="B2E7D7", fill_type="solid")
    elif color_mensaje.lower() == "rojo":
        ws[f"D{fila_actual}"].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    elif color_mensaje.lower() == "azul":
        ws[f"C{fila_actual}"].fill = PatternFill(start_color="AED3E3", end_color="AED3E3", fill_type="solid")
    else:
        ws[f"D{fila_actual}"].font = Font(color="000000")

    ws[f"C{fila_actual}"].alignment = center_alignment
    ws[f"D{fila_actual}"].alignment = center_alignment

    wb.save(archivo_excel)


def get_log_file():
    """Devuelve un archivo de log TXT único por ejecución."""
    fecha = datetime.now().strftime("%Y-%m-%d")
    hora = datetime.now().strftime("%H%M%S")  # misma lógica que HORA_EXCEL
    log_path = os.path.join(REPORTS_DIR, f"registro_pruebas_{fecha}_{hora}.txt")
    return open(log_path, "a", encoding="utf-8")