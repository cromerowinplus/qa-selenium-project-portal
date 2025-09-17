from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from utils.log_utils import log, log_excel,get_log_file
from utils.driver_factory import inicializar_driver

import locale
import time
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime, timedelta
import os

#Variables globales
#calendarioayer = datetime.now() - timedelta(days=1)
calendariohoy = datetime.now() 
calendariomañana = datetime.now() + timedelta(days=1)
horaexcel = datetime.now().strftime("%H%M%S")

def log(mensaje, archivo_log):
    hora = time.strftime("%Y-%m-%d %H:%M:%S")
    archivo_log.write(f"[{hora}] {mensaje}\n")
    archivo_log.flush()


def log_excel(nombre, mensaje="", negrita=False, color_mensaje="black"):
    ahora = datetime.now()
    fecha = ahora.strftime("%Y-%m-%d")
    hora = ahora.strftime("%H:%M:%S")
    
    archivo_excel = os.path.join("reports", f"registro_pruebas_{fecha}_{horaexcel}.xlsx")

    if not os.path.exists(archivo_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro"
        ws.append(["Fecha", "Hora", "Nombre", "Mensaje"])
        wb.save(archivo_excel)

    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    ws.append([fecha, hora, nombre, mensaje])

    fila_actual = ws.max_row

    # Configurar negrita y fondo gris en la columna C si 'negrita' es True
    if negrita:
        font_bold = Font(bold=True)
        fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris
        ws[f"C{fila_actual}"].font = font_bold
        ws[f"C{fila_actual}"].fill = fill_gray  # Solo aplicar gris en la columna C
    center_alignment = Alignment(horizontal='center', vertical='center')
    # Configurar color del mensaje en la columna D
    if color_mensaje.lower() == "verde":
        fill_verde = PatternFill(start_color="B2E7D7", end_color="B2E7D7", fill_type="solid")  # Verde
        ws[f"D{fila_actual}"].fill = fill_verde  # Solo aplicar verde en la columna C
    elif color_mensaje.lower() == "rojo":
        fill_rojo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Rojo
        ws[f"D{fila_actual}"].fill = fill_rojo  # Solo aplicar rojo en la columna C
        
    else:
        ws[f"D{fila_actual}"].font = Font(color="000000")  # Negro (por defecto)

    ws[f"C{fila_actual}"].alignment  = center_alignment  # Alinear
    ws[f"D{fila_actual}"].alignment  = center_alignment  # Alinear
    wb.save(archivo_excel)


def seleccionar_dia_calendario(driver, id_boton_datepicker, fecha_objetivo, logfile):
    wait = WebDriverWait(driver, 10)
    dia_objetivo = str(fecha_objetivo.day)

    # 1. Abrir el datepicker
    wait.until(EC.element_to_be_clickable((By.ID, id_boton_datepicker))).click()

    # 2. Esperar a que salgan días
    wait.until(lambda d: len(d.find_elements(By.XPATH, "//div[contains(@class,'mbsc-calendar-day-text')]")) >= 7)

    # 3. Buscar y clicar el día con un pequeño reintento
    for _ in range(3):
        dias = driver.find_elements(By.XPATH, "//div[contains(@class,'mbsc-calendar-day-text')]")
        for dia in dias:
            if dia.text.strip() == dia_objetivo:
                try:
                    driver.execute_script("arguments[0].scrollIntoView(true);", dia)
                    dia.click()
                    log(f"Clic en el día {dia_objetivo}", logfile)
                    return
                except Exception:
                    time.sleep(0.3)  # esperar y reintentar


def inicializar_driver():
    options = Options()
    prefs = {"profile.default_content_setting_values.geolocation": 1}
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--incognito")  
    service = Service(ChromeDriverManager().install())  # << Automático
    return webdriver.Chrome(service=service, options=options)


def clickatras(driver, log_file):
    try:
        boton = driver.find_element(By.ID, "tituloback")
        driver.execute_script("document.getElementById('tituloback').click();")
        log("Se hizo clic en el botón 'Ir atrás'", log_file)
    except Exception as e:
        log(f"No se pudo hacer clic en 'Ir atrás': {e}", log_file)


def entramodulo(driver, modulo):
    driver.execute_script(modulo)

def acceder_portal(driver, log_file):
    wait = WebDriverWait(driver, 8)
    driver.get("http://localhost/WinPlusPortal/web/")
    print("Página cargada correctamente.")

    log("Abierta la página principal", log_file)
    #log_excel("Personal", "Abierta la página principal", negrita=True)

    # Login
    campo_usuario = wait.until(EC.visibility_of_element_located((By.ID, "username_desktop")))
    campo_contraseña = driver.find_element(By.ID, "password_desktop")
    campo_usuario.send_keys("personal@IESTCRISTOBAL")
    campo_contraseña.send_keys("")

    boton_login = driver.find_element(By.XPATH, '//a[@onclick="login_presencia()"]')
    boton_login.click()

    # Esperar a que el botón de derechos esté visible y clicar
    boton_derechos = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[@onclick="login_presencia_derechos()"]'))
    )
    boton_derechos.click()

    log_excel("Derechos.Tipo=1 - Login", "Ok", negrita=True,color_mensaje="verde")
    log("Derechos.Tipo=1 - Login", log_file)

    log("Login realizado", log_file)
    log_excel("HistóricoUsuarios 0", "Ok", negrita=True,color_mensaje="verde")

    time.sleep(2)  # esperar login


def fichar(driver, log_file):
    entramodulo(driver, "pagefichar()")
    time.sleep(2)
    wait = WebDriverWait(driver, 2)
    try:
        fichaje_2 = driver.find_element(By.XPATH, "//a[@onclick='fichaje(2)']")
        if fichaje_2.is_displayed():
            fichaje_2.click()
            log("Se clicó fichaje salida", log_file)
        else:
            raise Exception("fichaje(2) no visible")
    except:
        boton_expandir = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "div#div_lfichar_secundario > h2 > a.ui-collapsible-heading-toggle")
        ))
        boton_expandir.click()
        time.sleep(1)
        fichaje_1 = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@onclick='fichaje(1)']")))
        fichaje_1.click()
        log("Se clicó entrada", log_file)
    
    try:
        popup = WebDriverWait(driver, 5).until(EC.any_of(
            EC.visibility_of_element_located((By.ID, "msgbox")),
            EC.visibility_of_element_located((By.ID, "msgerr"))
        ))
        if popup.get_attribute("id") == "msgbox":
            texto = driver.find_element(By.ID, "msgboxtext").text
            log_excel("Fichar", "Ok", negrita=True,color_mensaje="verde")
            log(f"Mensaje de éxito: {texto}", log_file)
        else:
            texto = driver.find_element(By.ID, "msgerrtext").text
            log_excel("Fichar", f"Error: {texto}", negrita=True,color_mensaje="rojo")
            log(f"Mensaje de error: {texto}", log_file)
        popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()
    except Exception as e:
        log(f"No apareció popup o hubo error al manejarlo: {e}", log_file)

    clickatras(driver, log_file)


def consultar_informacion(driver, log_file):
    entramodulo(driver, "pageconsulta()")
    time.sleep(1)

    #ayer = datetime.today() 
    #mañana = datetime.today() + timedelta(days=1)
    #fecha_ayer = ayer.strftime("%d-%m-%Y")
    #fecha_mañana = mañana.strftime("%d-%m-%Y")

    seleccionar_dia_calendario(driver, "fficdesde-button", calendariohoy,log_file)
    seleccionar_dia_calendario(driver, "ffichasta-button", calendariomañana,log_file)

    #driver.execute_script(f"document.getElementById('fficdesde').value = '{fecha_ayer}';")
    #driver.execute_script(f"document.getElementById('ffichasta').value = '{fecha_mañana}';")
    log(f"Rango de fechas: {calendariohoy} a {calendariomañana}", log_file)

    driver.find_element(By.XPATH, '//button[@onclick="getfichajes()"]').click()

    try:
        xpath = "//span[contains(@class, 'tableC') and contains(@class, 'c1') and (contains(@class, 'f1') or contains(@class, 'f2'))]"
        elemento = WebDriverWait(driver, 4).until(EC.presence_of_element_located((By.XPATH, xpath)))
        fecha_str = elemento.text.split()[0]
        fecha_span = datetime.strptime(fecha_str, "%d/%m/%Y").date()
        hoy = datetime.today().date()
        if fecha_span == hoy:
            log("La fecha del fichaje coincide con hoy", log_file)
            log_excel("Informacion", "Ok", negrita=True,color_mensaje="verde")
        else:
            log(f"Fecha NO coincide. Esperado {hoy}, encontrado {fecha_span}", log_file)
            log_excel("Informacion", f"Error: {fecha_span}", negrita=True,color_mensaje="rojo")
    except Exception as e:
        log(f"Error al buscar fichajes: {e}", log_file)
        log_excel("Informacion", "No hay fichajes", negrita=True,color_mensaje="rojo")

    clickatras(driver, log_file)

def consultar_acumulados(driver, log_file):
    entramodulo(driver, "pageacumulados('false')")
    time.sleep(1)

    #ayer = datetime.today() 
    #mañana = datetime.today() + timedelta(days=1)
    #fecha_ayer = ayer.strftime("%d-%m-%Y")
    #fecha_mañana = mañana.strftime("%d-%m-%Y")
    seleccionar_dia_calendario(driver, "facumdesde-button", calendariohoy,log_file)
    time.sleep(1)
    seleccionar_dia_calendario(driver, "facumhasta-button", calendariomañana,log_file)
    
    #driver.execute_script(f"document.getElementById('facumdesde').value = '{fecha_ayer}';")
    #driver.execute_script(f"document.getElementById('facumhasta').value = '{fecha_mañana}';")
    log(f"Acumulados: {calendariohoy} a {calendariomañana}", log_file)

    driver.find_element(By.XPATH, "//button[@onclick=\"getacumulados('false')\"]").click()
    time.sleep(2)
    try:
        xpath = "//li//span[@class='ui-li-count ui-body-inherit' or @class='ui-li-count red ui-body-inherit']"
        WebDriverWait(driver, 10).until(lambda d: len(d.find_elements(By.XPATH, xpath)) >= 1)
        resultados = [s for s in driver.find_elements(By.XPATH, xpath) if s.get_attribute("id") != "num_doc_total"]
        log(f"Se encontraron {len(resultados)} resultados visuales", log_file)
        if resultados:
            log_excel("Acumulados", "Ok", negrita=True,color_mensaje="Verde")
        else:
            log_excel("Acumulados", "Error - No hay acumulados", negrita=True,color_mensaje="Rojo")
    except Exception as e:
        log(f"No se encontraron acumulados: {e}", log_file)
        log_excel("Acumulados", "No se detectaron resultados", negrita=True,color_mensaje="Rojo")

    clickatras(driver, log_file)

def validacion(driver, log_file):
    entramodulo(driver,"pagevalidacion()")
    log("Entrando en Validacion", log_file)

    time.sleep(1)

    # Calcular fechas
    hoy = datetime.today().strftime("%d-%m-%Y")
    ayer = (datetime.today() - timedelta(days=1)).strftime("%d-%m-%Y")

    seleccionar_dia_calendario(driver, "fval-button", calendariomañana,log_file)
    time.sleep(1)
    seleccionar_dia_calendario(driver, "fval-button", calendariohoy,log_file)

    # Paso 1: Establecer fecha de ayer
    #driver.execute_script(f"""
    #    const input = document.getElementById('fval');
    #    input.value = '{ayer}';
    #    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
    #""")
    log(f"Fecha establecida en Validación: {ayer}", log_file)
    time.sleep(1)

    # Paso 2: Establecer fecha de hoy
    #driver.execute_script(f"""
    #    const input = document.getElementById('fval');
    #    input.value = '{hoy}';
    #    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
    #""")
    log(f"Fecha actualizada en Validación: {hoy}", log_file)
    time.sleep(1)

    log_excel("Validación", "", negrita=True)

    # Paso 3: Click en pestaña 'Fichajes'
    driver.execute_script("tabval('tvalfichajes')")
    log("Clic en pestaña Fichajes", log_file)
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tvalfichajes .tableA.c1"))
        )
        log("Validación - Fichajes: tabla encontrada", log_file)
        log_excel("Validación - Fichajes", "Ok",color_mensaje="verde")
    except Exception as e:
        log(f"Validación - Fichajes: tabla NO encontrada. {e}", log_file)
        log_excel("Validación - Fichajes", "Tabla no encontrada",color_mensaje="rojo")

    # Paso 4: Click en pestaña 'Contadores'
    driver.execute_script("tabval('tvalcontadores')")
    log("Clic en pestaña Contadores", log_file)
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tvalcontadores .tableC.c1"))
        )
        log("Validación - Contadores: tabla encontrada", log_file)
        log_excel("Validación - Contadores", "Ok",color_mensaje="verde")
    except Exception as e:
        log(f"Validación - Contadores: tabla NO encontrada. {e}", log_file)
        log_excel("Validación - Contadores", "Tabla no encontrada",color_mensaje="rojo")
        
    # Paso 4: Click en pestaña 'Incidencias'
    driver.execute_script("tabval('tvalincidencias')")
    log("Clic en pestaña Incidencias", log_file)
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tvalincidencias .tableA.c1"))
        )
        log("Validación - Incidencias: tabla encontrada", log_file)
        log_excel("Validación - Incidencias", "Ok",color_mensaje="verde")
    except Exception as e:
        log(f"Validación - Incidencias: tabla NO encontrada. {e}", log_file)
        log_excel("Validación - Incidencias", "Tabla no encontrada",color_mensaje="rojo")


    clickatras(driver,log_file)

def solicitar_fichaje_manual(driver, log_file):
    ayer = (datetime.today()).strftime("%d-%m-%Y")
    mañana = (datetime.today() + timedelta(days=1)).strftime("%d-%m-%Y")
    # Entrar a Fichajes Validación
    entramodulo(driver, "pagefichajesval('false')")
    log("Entrando en FichajesValidación", log_file)
    log_excel("Fichajes de validacion","",negrita=True)
    time.sleep(1)

    seleccionar_dia_calendario(driver, "fvaldesde-button", calendariohoy,log_file)
    time.sleep(1)
    seleccionar_dia_calendario(driver, "fvalhasta-button", calendariomañana,log_file)
    #driver.execute_script(f"document.getElementById('fvaldesde').value = '{ayer}';")
    #driver.execute_script(f"document.getElementById('fvalhasta').value = '{mañana}';")
    driver.find_element(By.XPATH, '//button[@onclick="getfvalidacion()"]').click()

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tfichajesval span.tableC.c3"))
        )
        spans = driver.find_elements(By.CSS_SELECTOR, "#tfichajesval span.tableC.c3, .tableC.c4, .tableA.c3, .tableA.c4")
        con_texto = [s for s in spans if s.text.strip()]
        for i, s in enumerate(con_texto, start=1):
            log(f"FichajesValidación: Span {i} = '{s.text.strip()}'", log_file)
        if con_texto:
            log_excel("Fichajes validación - Consulta", "Ok",color_mensaje="verde")
        else:
            log_excel("Fichajes validación - Consulta", "Sin datos", negrita=False,color_mensaje="rojo")
    except Exception as e:
        log(f"Error al buscar datos en tfichajesval: {e}", log_file)
        log_excel("Fichajes validación - Consulta", "No se han encontrado datos", negrita=False,color_mensaje="rojo")

    # Crear nuevo fichaje manual
    driver.find_element(By.XPATH, '//button[@onclick="gotofichajemanual()"]').click()

    seleccionar_dia_calendario(driver, "fm_fecha-button", calendariohoy,log_file)
    #driver.execute_script(f"document.getElementById('fm_fecha').value = '{ayer}';")
    #driver.execute_script(f"document.getElementById('fm_hora').value = '10:00';")
    clicabyid("fm_hora-button",log_file)
    clicabyclase("ok",log_file)
    escribir_observacion("fm_notas","Observacion escrita desde selenium",log_file)

    wait = WebDriverWait(driver, 3)
    boton_solicitar = wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@onclick="fm_solicitar_fichaje(1)"]')))
    boton_solicitar.click()
    log("Solicitamos fichaje entrada", log_file)

    time.sleep(1)
    driver.find_element(By.XPATH, '//button[@onclick="dialogbox_response(1)"]').click()
    log("Clicamos OK en el diálogo de confirmación", log_file)

    try:
        popup = WebDriverWait(driver, 5).until(EC.any_of(
            EC.visibility_of_element_located((By.ID, "msgbox")),
            EC.visibility_of_element_located((By.ID, "msgerr"))
        ))
        if popup.get_attribute("id") == "msgbox":
            texto = driver.find_element(By.ID, "msgboxtext").text
            log(f"Popup éxito fichaje manual: {texto}", log_file)
            log_excel("Fichaje manual", "Ok",color_mensaje="verde")
            wait = WebDriverWait(driver, 3)
            wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#msgbox a[data-rel='back']"))).click()
            driver.find_element(By.XPATH, '//button[@onclick="gotofichajemanual()"]').click()
            log("Volvemos a la pantalla de fichaje manual tras éxito", log_file)
        else:
            texto = driver.find_element(By.ID, "msgerrtext").text
            log(f"Popup error fichaje manual: {texto}", log_file)
            log_excel("Fichaje manual", f"Error: {texto}", negrita=True,color_mensaje="rojo")
            popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()
    except Exception as e:
        log(f"No apareció popup o falló el manejo: {e}", log_file)
        log_excel("Fichaje manual", "Error con el popup", negrita=True,color_mensaje="rojo")

    

    # Esperar que el campo esté visible/interactivo
    campo_fecha = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.ID, "fm_fecha"))
    )

    # Cambiar a mañana, simular cambio
    driver.execute_script(f"arguments[0].value = '{mañana}';", campo_fecha)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", campo_fecha)
    campo_fecha.send_keys(Keys.ENTER)
    time.sleep(1)
    
    # Volver a ayer, simular cambio
    driver.execute_script(f"arguments[0].value = '{ayer}';", campo_fecha)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", campo_fecha)
    campo_fecha.send_keys(Keys.ENTER)
    time.sleep(1)
    
    try:
        elementos = driver.find_elements(By.CSS_SELECTOR, "#fm_solicitudes li")
        if elementos:
            log("Solicitud detectada en listado", log_file)
            log_excel("Fichaje manual - Existe Solicitud Creada", "Ok",color_mensaje="verde")
        else:
            log("No se detectó ninguna solicitud", log_file)
            log_excel("Fichaje manual - Existe Solicitud Creada", "Error - No detecta solicitud",color_mensaje="rojo")
    except Exception as e:
        log(f"Error al comprobar solicitudes: {e}", log_file)
        log_excel("Fichaje manual - Solicitud Creada", "Error al comprobar", negrita=True,color_mensaje="rojo")

    clickatras(driver,log_file)
    time.sleep(1)
    clickatras(driver,log_file)
    time.sleep(1)
    clickatras(driver,log_file)
    time.sleep(1)
    clickatras(driver,log_file)

def usar_tarjeta(driver, log_file):
    entramodulo(driver, "usartarjeta()")
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.ID, "personal_qrcode"))
        )
        log_excel("Uso de tarjeta", "Ok",negrita=True,color_mensaje="verde")
        log("Uso de tarjeta Ok", log_file)
    except TimeoutException:
        log_excel("Uso de tarjeta", "Error - QR no visible", negrita=True,color_mensaje="rojo")
        log("Uso de tarjeta ERROR", log_file)
    clicarhome(driver,log_file)

def clicabyid(id,log_file):
    driver.find_element(By.ID, id).click()
    log(f"Clicamos sobre {id} ",log_file)

def clicabyclase( clase, log_file, timeout=5):
    try:
        boton = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, f".mdtimepicker:not(.hidden) .mdtp__button.{clase}")
            )
        )
        boton.click()
        log(f"Clicamos sobre {clase}", log_file)
    except Exception as e:
        msg = f"Error al intentar clicar {clase}: {e}"
        log(msg, log_file)        
        log_excel(msg,log_file,color_mensaje="rojo")             
        raise



def clicar_segundo_boton_combobox(driver, log_file):
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.ID, "combobox"))
        )
        botones = driver.find_elements(By.CSS_SELECTOR, "#combobox button")
        
        # Obtener el texto del botón
        texto_boton = botones[1].text.strip()
        
        #Clicamos el boton
        botones[1].click()

        if log_file: log(f"Clicado la incidencia/causa {texto_boton}", log_file)

        return texto_boton
    except Exception as e:
        if log_file: log(f"Error clicando el combobox: {e}", log_file)
        raise

def seleccionar_opcion_combobox(driver, log_file, texto_buscado, timeout=10):
    wait = WebDriverWait(driver, timeout)

    # --- MULTISELECT ---
    try:
        popup = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#combobox_multiselect")))
        checks = popup.find_elements(By.CSS_SELECTOR, "#comboboxmulti_fieldset .ui-checkbox")
        for cont in checks:
            label = cont.find_element(By.TAG_NAME, "label")
            texto = (label.text or "").strip()
            if texto == texto_buscado:
                label.click()
                # OK: espera + JS click, pero NO rompas si falla
                try:
                    ok = wait.until(EC.element_to_be_clickable((By.ID, "multiselect_ok")))
                    driver.execute_script("arguments[0].click();", ok)
                except Exception as e:
                    if log_file: log(f"[MULTISELECT] Seleccionada '{texto}' (OK pendiente): {e}", log_file)
                else:
                    if log_file: log(f"[MULTISELECT] Seleccionada '{texto}'", log_file)
                return texto
    except Exception:
        pass  # Si no hay multiselect, sigue

    # --- SIMPLE ---
    try:
        wait.until(EC.visibility_of_any_elements_located((By.CSS_SELECTOR, "#combobox button, .ui-controlgroup-controls button")))
        for btn in driver.find_elements(By.CSS_SELECTOR, "#combobox button, .ui-controlgroup-controls button"):
            texto = (btn.text or btn.get_attribute("innerText") or "").strip()
            if texto == texto_buscado:
                btn.click()
                if log_file: log(f"[SIMPLE] Seleccionada '{texto}'", log_file)
                return texto
    except Exception:
        pass

    if log_file: log(f"No se encontró la opción '{texto_buscado}' en el combobox", log_file)
    return None


def escribir_observacion(identificador, texto, log_file):
    try:
        # Esperar a que el campo esté presente y visible
        campo = WebDriverWait(driver, 3).until(
            EC.visibility_of_element_located((By.ID, identificador))
        )
  
        # Limpiar el campo antes de escribir
        campo.clear()

        # Escribir el texto en el campo
        campo.send_keys(f"{texto}")

        # Registro en el log
        log("Se ha escrito en el campo", log_file)

    except Exception as e:
        log(f"Error al escribir en el campo {identificador}: {e}", log_file)

def click_onclick(driver, tipo,onclick_value, log_file, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, f'//{tipo}[@onclick="{onclick_value}"]'))
        )
        element.click()
        log(f'Click en botón con onclick="{onclick_value}"', log_file)
        return True
    except Exception as e:
        log(f'No se pudo hacer click en botón con onclick="{onclick_value}" -> {e}', log_file)
        return False

def guardarsolicitud():
            # Clicamos boton guardar)
    xpath_mas_selectivo = "(//button[@onclick='savesolicitud()' and not(@disabled) and " \
                          "not(contains(@class,'ui-state-disabled')) and " \
                          "not(contains(@style,'display: none')) and " \
                          "not(contains(@style,'visibility: hidden'))])[last()]"
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_mas_selectivo))).click()
    #Clicamos en el popup de "Si"
    clicar_si(driver,"button","dialogbox_response(1)",log_file,"Solicitudes- Nueva")


def clicar_si(driver,tipo,identificador,log_file,mensajepersonalizado):
    click_onclick(driver,tipo,identificador,log_file)
    try:
        popup = WebDriverWait(driver, 5).until(EC.any_of(
            EC.visibility_of_element_located((By.ID, "msgbox")),
            EC.visibility_of_element_located((By.ID, "msgerr"))
        ))
        if popup.get_attribute("id") == "msgbox":
            texto = driver.find_element(By.ID, "msgboxtext").text
            log_excel(mensajepersonalizado, "Ok", negrita=False,color_mensaje="verde")
            log(f"Mensaje de éxito: {texto}", log_file)
        else:
            texto = driver.find_element(By.ID, "msgerrtext").text
            log_excel(mensajepersonalizado, f"Error: {texto}", negrita=False,color_mensaje="rojo")
            log(f"Mensaje de error: {texto}", log_file)
            clickatras(driver,log_file)
            time.sleep(0.7)
            clickatras(driver,log_file)
        popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()


    except Exception as e:
        log(f"No apareció popup o hubo error al manejarlo: {e}", log_file)

def compruebatopes(driver, log_file):
    time.sleep(1)
    clicabyid("solicitudes_topes_personal",log_file)
    time.sleep(1)
    #wait = WebDriverWait(driver, 10)
    #wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@onclick="vertopes_personal()"]'))).click()
    log("Entramos en topes",log_file)
    if len(driver.find_elements(By.ID, "lista_contadores_topados")) > 0:
        log("Existen topes'", log_file)
        log_excel("Solicitudes- Topes","ok",negrita=False,color_mensaje="verde")
    else:
        log("Error, no existen los topes", log_file)
        log_excel("Solicitudes- Topes","Error No existen topes",negrita=True,color_mensaje="rojo")

    log("salimos de los topes",log_file)
    time.sleep(1)
     
    clickatras(driver,log_file)



def creasolicitud(driver,log_file):
    global textoincidenciasolicitud, textocausasolicitud 
    click_onclick(driver,"button","versolicitud(-1)",log_file)


    # Clic en el campo "desde" → seleccionar AYER
    seleccionar_dia_calendario(driver, "sol_fdesde-button", calendariohoy,log_file)
    time.sleep(1)
    # Clic en el campo "hasta" → seleccionar MAÑANA
    seleccionar_dia_calendario(driver, "sol_fhasta-button", calendariomañana,log_file)

    #Clicamos las horas


    clicabyid("sol_hfin-button",log_file)
    clicabyclase("ok",log_file)

    clicabyid("sol_hinicio-button",log_file)
    clicabyclase("ok",log_file)

    #getElementById("cbx_fsol_incidencia").click();

    clicabyid("cbx_sol_incidencia",log_file)
    textoincidenciasolicitud=clicar_segundo_boton_combobox(driver,log_file)
    clicabyid("cbx_sol_causa",log_file)
    textocausasolicitud=clicar_segundo_boton_combobox(driver,log_file)
    escribir_observacion("sol_notas","Observacion escrita desde Selenium",log_file)

    guardarsolicitud()

def existe_solicitud(driver, log_file, timeout=10):
    desde = calendariohoy.strftime("%d-%m-%Y"); hasta = calendariomañana.strftime("%d-%m-%Y")
    xp = f'//ul[@id="lsolicitudes"]//a[.//div[normalize-space()="Desde: {desde} Hasta: {hasta}"]]'
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
    if log_file: log(f"OK: existe solicitud con fechas {desde} - {hasta}", log_file)
    return True

def busca_solicitud(driver,log_file):
    seleccionar_dia_calendario(driver, "fsoldesde-button", calendariohoy,log_file)
    time.sleep(1)
    # Clic en el campo "hasta" → seleccionar MAÑANA
    seleccionar_dia_calendario(driver, "fsolhasta-button", calendariomañana,log_file)


    clicabyid("cbx_fsol_incidencia",log_file)
    #seleccionar_opcion_combobox(driver, log_file, textoincidenciasolicitud)
    seleccionar_opcion_combobox(driver, log_file, "Adelanto en la entrada")
    
    #click_onclick(driver,"button","combobox_multi_aceptar('')",log_file)
    driver.execute_script("document.getElementById('multiselect_ok').click();")


    clicabyid("cbx_fsol_solicitud",log_file)

    #seleccionar_opcion_combobox(driver, log_file, textocausasolicitud)
    seleccionar_opcion_combobox(driver,log_file,"Horas Extra 1")

    driver.execute_script("document.getElementById('multiselect_ok').click();")
    wait = WebDriverWait(driver, 10)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@onclick="getsol_personal()"]'))).click()

    if existe_solicitud(driver,log_file):
        log("Existe solicitud",log_file)
        log_excel("Solicitudes Buscar", "Ok", negrita=False,color_mensaje="verde")
    else:
        log_excel("Solicitudes Buscar", "Error, no existe", negrita=True,color_mensaje="rojo")
        log("No existe solicitud",log_file)

    clickatras(driver,log_file)
    clickatras(driver,log_file)

def seleccionar_dia_calendarioOLD(driver, numero_dia, timeout=10):
    wait = WebDriverWait(driver, timeout)
    dia = wait.until(EC.element_to_be_clickable(
        (By.XPATH, f"//a[@class='dbEvent ui-btn ui-mini m0 ui-btn-a' or @class='dbEvent ui-btn ui-mini m0 ui-btn-active'][normalize-space(text())='{numero_dia}']")
    ))
    dia.click()

def solicitudes(driver, log_file):

    entramodulo(driver, "pagesolicitudes()")
    log_excel("Solicitudes", "", negrita=True)

    #Comprobamos si existen topes
    compruebatopes(driver,log_file)

    creasolicitud(driver,log_file)

    busca_solicitud(driver,log_file)

def set_fecha(driver,identificador, fecha , log_file):
    wait = WebDriverWait(driver, 10)

    fdesde = wait.until(EC.visibility_of_element_located((By.ID, identificador)))
    fdesde.send_keys(fecha)
    fdesde.send_keys(Keys.TAB)  # dispara onchange
    log(f"Escribimos la fecha: {fecha} en {identificador}",log_file)

def guardar_cambio(driver,log_file):

    driver.execute_script("arguments[0].click();", WebDriverWait(driver,10).until(lambda d: next(e for e in d.find_elements(By.XPATH,'//button[starts-with(@onclick,"cp_save")]') if e.is_displayed() and e.is_enabled())))

    log("Clicamos Guardar",log_file)
    clicar_si(driver,"button","dialogbox_response(1)",log_file,"Gestion- Añadir")
    
def añadir(driver,log_file):
    global horariocambiohorario
    click_onclick(driver,"a","cp_ver(-1)",log_file)
    #Añadimos las fechas a la fuerza porque el calendario es el viejo, lo van a cambiar al nuevo
    set_fecha(driver,"cp_fdesde_edit",calendariohoy.strftime("%d-%m-%Y"),log_file)
    set_fecha(driver,"cp_fhasta_edit",calendariomañana.strftime("%d-%m-%Y"),log_file)

    clicabyid("cbx_cp_horario_edit",log_file)
    horariocambiohorario=clicar_segundo_boton_combobox(driver,log_file)
    
    guardar_cambio(driver,log_file)
    
def gestionar(driver,log_file):
    click_onclick(driver,"a","gotonext('#page_cambios_personal_filtro');",log_file)

    set_fecha(driver,"cp_fdesde",calendariohoy.strftime("%d-%m-%Y"),log_file)
    set_fecha(driver,"cp_fhasta",calendariomañana.strftime("%d-%m-%Y"),log_file)

    clicabyid("cbx_cp_horario",log_file)
    #seleccionar_horario_combobox(driver,"ANA MARIN (9,3 a 16,3)")
    seleccionar_horario_combobox(driver,horariocambiohorario)
    click_onclick(driver,"button","cp_lista()",log_file)
    existe_horario(driver,log_file)
    clickatras(driver,log_file)
    time.sleep(0.5)
    clickatras(driver,log_file)

def existe_horario(driver,log_file, timeout=5):
    try:
        wait = WebDriverWait(driver, timeout)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#cp_lista > li")))
        log("Existe horario",log_file)
        log_excel("Gestion - Gestionar", "Ok",color_mensaje="verde")
    except TimeoutException:
        log("No existe horario",log_file)
        log_excel("Gestion - Gestionar", "No existe horario",negrita=False,color_mensaje="rojo")

def seleccionar_horario_combobox(driver, texto, timeout: int = 10) -> bool:
    wait = WebDriverWait(driver, timeout)

    # 1) Asegura que el popup está visible
    wait.until(EC.visibility_of_element_located((By.ID, "combobox")))

    # 2) Busca el botón cuyo texto (normalizado) coincide
    def _find_btn(drv):
        norm = lambda s: " ".join((s or "").split()).strip()
        objetivo = norm(texto)
        for b in drv.find_elements(By.CSS_SELECTOR, '#combobox button.ui-btn'):
            if b.is_displayed() and b.is_enabled() and 'ui-state-disabled' not in (b.get_attribute('class') or ''):
                if norm(b.text) == objetivo:  # coincidencia exacta normalizada
                    return b
        # Fallback: coincidencia parcial (por si hay mayúsculas/minúsculas)
        for b in drv.find_elements(By.CSS_SELECTOR, '#combobox button.ui-btn'):
            if b.is_displayed() and objetivo.lower() in norm(b.text).lower():
                return b
        return False  # sigue esperando

    btn = wait.until(_find_btn)
    # 3) Click robusto
    driver.execute_script("arguments[0].click();", btn)

def comprobar_informe(driver,log_file,nombretitulo):
    ventana_principal = driver.current_window_handle
    manejadores_antes = driver.window_handles
    try:
        # Esperar a que aparezca una nueva ventana/pestaña (máximo 10 segundos)
        WebDriverWait(driver, 4).until(EC.new_window_is_opened(manejadores_antes))
        
        # Obtener todos los manejadores disponibles
        todos_manejadores = driver.window_handles
        
        # Encontrar el nuevo manejador
        nuevo_manejador = [manejador for manejador in todos_manejadores if manejador not in manejadores_antes][0]
        
        # Cambiar a la nueva ventana
        driver.switch_to.window(nuevo_manejador)
        driver.close()

        driver.switch_to.window(ventana_principal)
        log("Se ha abierto el informe",log_file)
        log_excel(f"{nombretitulo}","Ok",negrita=False,color_mensaje="verde")
    
    except:
        log("No se ha abierto el informe",log_file)
        log_excel(f"{nombretitulo}","No se abre el informe",negrita=True,color_mensaje="rojo")
        

def imprimir(driver,log_file,nombretitulo):
    click_onclick(driver,"a","calp_print()",log_file)
    boton_imprimir = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, "ul.ui-listview.ui-listview-inset.ui-corner-all.ui-shadow a[onclick='calp_report_print()']"))
    )
    boton_imprimir.click()
    
    comprobar_informe(driver,log_file,nombretitulo)
    clickatras(driver,log_file)

def planificacion(driver,log_file):
    log_excel("Planificacion de Turnos","ESTO HAY QUE COMPROBARLO MANUAL",negrita=True,color_mensaje="rojo")
    imprimir(driver,log_file,"Planificacion de turnos Imprimir")

def calendario(driver,log_file):
    log_excel("Calendario Personal","ESTO HAY QUE COMPROBARLO MANUAL",negrita=True,color_mensaje="rojo")
    imprimir(driver,log_file,"Calendario Personal Imprimir")
    # estoy poniendo los nombres que voy a enviar al log

def gestion(driver,log_file):
    log_excel("Cambios Horarios", "", negrita=True)
    añadir(driver,log_file)
    gestionar(driver,log_file)

def cambios_horarios(driver,log_file):
    entramodulo(driver, "cp_page(true)")
    gestion(driver,log_file)
    calendario(driver,log_file)
    planificacion(driver,log_file)
    clickatras(driver,log_file)

def ficharproduccion(driver,log_file):
    global proyecto 
    proyecto = "Proyecto 0000"
    entramodulo(driver, "pagefichar()")
    time.sleep(2)
    wait = WebDriverWait(driver, 2)

    boton_expandir = wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "div#div_lfichar_secundario > h2 > a.ui-collapsible-heading-toggle")
    ))
    boton_expandir.click()
    time.sleep(1)
    click_onclick(driver,"a","centrosdecoste()",log_file)

    click_onclick(driver,"a","combobox('cbx_costes_linea1','seleccione_costes_linea1',combobox_costes_linea1)",log_file)
    seleccionar_proyecto_0000(driver)

    #Si existe para poner la cantidad
    try:
        driver.find_element(By.ID, 'cc_cantidad_multilinea')
        escribir_observacion("cc_cantidad_multilinea", 99, log_file)
    except NoSuchElementException:
        pass
    time.sleep(1)
    click_onclick(driver,"a","fichajecentro('000000000')",log_file)

    
    try:
        popup = WebDriverWait(driver, 5).until(EC.any_of(
            EC.visibility_of_element_located((By.ID, "msgbox")),
            EC.visibility_of_element_located((By.ID, "msgerr"))
        ))
        if popup.get_attribute("id") == "msgbox":
            texto = driver.find_element(By.ID, "msgboxtext").text
            log_excel("Fichar", "Ok", negrita=False,color_mensaje="verde")
            log(f"Mensaje de éxito: {texto}", log_file)
        else:
            texto = driver.find_element(By.ID, "msgerrtext").text
            log_excel("Fichar", f"Error: {texto}", negrita=True,color_mensaje="rojo")
            log(f"Mensaje de error: {texto}", log_file)
        popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()
    except Exception as e:
        log(f"No apareció popup o hubo error al manejarlo: {e}", log_file)
    time.sleep(241)
    click_onclick(driver,"a","centrosdecoste()",log_file)
    
    escribir_observacion("cc_cantidad_multilinea", 99, log_file)
    #Esperamos 60 segundos porque si no tendremos fichaje duplicado
    
    click_onclick(driver,"a","fichajecentro('000000000')",log_file)
    try:
        popup = WebDriverWait(driver, 5).until(EC.any_of(
            EC.visibility_of_element_located((By.ID, "msgbox")),
            EC.visibility_of_element_located((By.ID, "msgerr"))
        ))
        if popup.get_attribute("id") == "msgbox":
            texto = driver.find_element(By.ID, "msgboxtext").text
            log_excel("Fichar", "Ok", negrita=False,color_mensaje="verde")
            log(f"Mensaje de éxito: {texto}", log_file)
        else:
            texto = driver.find_element(By.ID, "msgerrtext").text
            log_excel("Fichar", f"Error: {texto}", negrita=True,color_mensaje="rojo")
            log(f"Mensaje de error: {texto}", log_file)
            popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()
            click_onclick(driver,"a","centrosdecoste()",log_file)
            escribir_observacion("cc_cantidad_multilinea", 99, log_file)
            #Esperamos 60 segundos porque si no tendremos fichaje duplicado
            
            click_onclick(driver,"a","fichajecentro('000000000')",log_file)
        popup.find_element(By.CSS_SELECTOR, "a[data-rel='back']").click()
    except Exception as e:
        log(f"No apareció popup o hubo error al manejarlo: {e}", log_file)
    
def clicarhome(driver,log_file):
    logo = driver.find_element(By.CSS_SELECTOR, "#pageqrpersonal img.logo_app")
    logo.click()

def seleccionar_proyecto_0000(driver):
    try:
        # Esperar hasta que el botón con el texto "Proyecto 0000" sea visible
        boton_proyecto_0000 = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//button[normalize-space(text())='Proyecto 0000']"))
        )
        
        # Hacer clic en el botón
        boton_proyecto_0000.click()
        
    
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def centrosdecoste(driver,log_file):
    ficharproduccion(driver,log_file)

driver = inicializar_driver()
with get_log_file() as log_file:
    try:
        acceder_portal(driver, log_file)
        #fichar(driver, log_file)
        #consultar_informacion(driver, log_file)
        #consultar_acumulados(driver, log_file)
        #validacion(driver, log_file)
        solicitar_fichaje_manual(driver, log_file)
        usar_tarjeta(driver, log_file)
        #solicitudes(driver, log_file)
        cambios_horarios(driver,log_file)
        #centrosdecoste(driver,log_file) ##Se puede fichar pero no funciona la consulta por un problema de oracle



    finally:
        input("Presiona Enter para cerrar el navegador...")
        driver.quit()
        log("Navegador cerrado y prueba finalizada", log_file)

